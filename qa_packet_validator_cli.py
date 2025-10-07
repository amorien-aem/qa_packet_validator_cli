
import fitz  # PyMuPDF
import csv
import re
import matplotlib.pyplot as plt
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.worksheet.table import Table, TableStyleInfo
import os
import argparse

def validate_pdf(pdf_path):
    base_name = os.path.splitext(os.path.basename(pdf_path))[0]
    output_dir = os.path.dirname(pdf_path)
    csv_path = os.path.join(output_dir, f"{base_name}_validation_summary.csv")
    excel_path = os.path.join(output_dir, f"{base_name}_validation_summary.xlsx")
    dashboard_path = os.path.join(output_dir, f"{base_name}_dashboard.png")

    REQUIRED_FIELDS = [
        "Customer Name", "Customer P.O. Number", "Customer Part Number",
        "Customer Part Number Revision", "OEM Part Number", "OEM Lot Number",
        "OEM Date Code", "OEM Cage Code", "AEM Part Number", "AEM Lot Number",
        "AEM Date Code", "AEM Cage Code", "Customer Quality Clauses",
        "FAI Form 3", "Solderability Test Report", "DPA", "Visual Inspection Record",
        "Shipment Quantity", "Reel Labels", "Certificate of Conformance", "Route Sheet",
        "Part Number", "Lot Number", "Date", "Resistance", "Dimension", "Test Result"
    ]

    NUMERICAL_RANGES = {
        "Resistance": (95, 105),
        "Dimension": (0.9, 1.1)
    }

    anomalies = []
    critical_issues = []
    field_presence = defaultdict(int)
    all_fields = []

    def extract_fields(text):
        fields = {}
        for field in REQUIRED_FIELDS:
            pattern = rf"{field}[:\s]*([^\n]+)"
            match = re.search(pattern, text, re.IGNORECASE)
            if match:
                fields[field] = match.group(1).strip()
        return fields

    def validate_numerical(field, value):
        try:
            val = float(re.findall(r"[\d.]+", value)[0])
            min_val, max_val = NUMERICAL_RANGES[field]
            return min_val <= val <= max_val
        except:
            return False

    def check_consistency(field_name):
        values = [fields.get(field_name) for fields in all_fields if field_name in fields]
        return len(set(values)) == 1

    doc = fitz.open(pdf_path)

    for page_num in range(len(doc)):
        page = doc.load_page(page_num)
        text = page.get_text()
        fields = extract_fields(text)
        all_fields.append(fields)

        for field in REQUIRED_FIELDS:
            if field not in fields:
                anomalies.append([page_num + 1, field, "Missing"])
            else:
                field_presence[field] += 1

        for field in NUMERICAL_RANGES:
            if field in fields and not validate_numerical(field, fields[field]):
                anomalies.append([page_num + 1, field, f"Out of range: {fields[field]}"])
                critical_issues.append([page_num + 1, field, fields[field]])

    for field in ["Part Number", "Lot Number", "Date"]:
        if not check_consistency(field):
            anomalies.append(["All Pages", field, "Inconsistent values"])
            critical_issues.append(["All Pages", field, "Inconsistent values"])

    with open(csv_path, "w", newline='') as f:
        writer = csv.writer(f)
        writer.writerow(["Page", "Field", "Issue"])
        writer.writerows(anomalies)

    wb = Workbook()
    ws = wb.active
    ws.title = "QA Anomalies"

    headers = ["Page", "Field", "Issue"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)

    for row_num, row_data in enumerate(anomalies, start=2):
        for col_num, cell_value in enumerate(row_data, start=1):
            ws.cell(row=row_num, column=col_num, value=cell_value)

    table_ref = f"A1:C{len(anomalies)+1}"
    table = Table(displayName="AnomalyTable", ref=table_ref)
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    table.tableStyleInfo = style
    ws.add_table(table)

    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_length + 2

    wb.save(excel_path)

    plt.figure(figsize=(12, 6))
    plt.bar(field_presence.keys(), field_presence.values(), color='skyblue')
    plt.title("Field Presence Across PDF Pages")
    plt.xlabel("Field Name")
    plt.ylabel("Number of Pages Present")
    plt.xticks(rotation=90)
    plt.tight_layout()
    plt.savefig(dashboard_path)

    return csv_path, excel_path, dashboard_path, len(anomalies), len(critical_issues)

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Validate a QA Packet PDF.")
    parser.add_argument("pdf_path", help="Path to the PDF file to validate")
    args = parser.parse_args()

    csv_path, excel_path, dashboard_path, anomaly_count, critical_count = validate_pdf(args.pdf_path)
    print(f"Validation complete.\nAnomalies: {anomaly_count}\nCritical Issues: {critical_count}")
    print(f"CSV saved to: {csv_path}")
    print(f"Excel saved to: {excel_path}")
    print(f"Dashboard saved to: {dashboard_path}")
